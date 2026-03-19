"""Liest den GoDi-Plan aus der Excel-Datei."""
from __future__ import annotations

import glob
import logging
import os
import re
import subprocess
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Optional

import openpyxl

from config import (GODI_PLAN_DIR, EXCEL_ROWS, ABKUENDIGUNGEN_ROWS,
                     EINLADUNG_ROW_START, EINLADUNG_ROW_END, EINLADUNG_COLS)

log = logging.getLogger(__name__)

# Farb-Codes für die automatische Erkennung (Excel fgColor RGB)
SONG_COLOR    = "FFC5E0B4"  # Grün  → Lied
PREDIGT_COLOR = "FFFFA7A7"  # Rot   → Predigt-Bibelstelle
LESUNG_COLOR  = "FFF8CBAD"  # Orange → Lesung-Bibelstelle

# Bekannte Lied-Präfixe für die Inhaltserkennung
_SONG_BOOKS = re.compile(
    r'^(FJ\d*|GLS|SUG|IWDD|Loben|SGIDH)\s+\d+',
    re.IGNORECASE
)
_SONG_PREFIXES = ("lobpreisstrophe:", "kinderlied:", "sonstige lieder")

# Muster für eine einzelne Bibelstelle (kein Semikolon = keine Liste)
_BIBLE_REF = re.compile(
    r'^((?:\d\.\s*)?[A-Za-zäöüÄÖÜß]+)\s+\d+',
)


@dataclass
class InvitationEvent:
    """Ein Termin für die 'Herzliche Einladung' Folie."""
    date_str: str = ""       # "Di 17.03.26"
    time_str: str = ""       # "19:00"
    event_name: str = ""     # "Gebetsstunde"
    note: str = ""           # "Abend-GoDi fällt aus!" (optional, aus Spalte F)


@dataclass
class SongEntry:
    """Ein Lied aus dem GoDi-Plan."""
    raw: str                    # Original-Zelltext
    category: str = ""          # "Gemeindelied", "Lobpreisstrophe", "Kinderlied", etc.
    book: str = ""              # "FJ1", "SUG", "GLS", etc.
    number: str = ""            # "235", "61", etc.
    title: str = ""             # Volltitel nach " - "
    title_words: list = field(default_factory=list)  # Erste 3 Wörter des Titels
    slot_key: str = ""          # "song1" bis "song7"


@dataclass
class GodiPlanData:
    """Alle Daten eines Gottesdienstes aus dem GoDi-Plan."""
    service_header: str = ""        # "Gottesdienst am 08.03.2026 (Okuli)"
    theme: str = ""                 # "Der Blick nach vorn"
    date_str: str = ""              # "08.03.2026"
    kirchenkalender: str = ""       # "Okuli"
    greeting_verse: str = ""        # Begrüßungsvers
    lesung_reference: str = ""      # "Lukas 9, 57-62"
    predigt1_reference: str = ""    # "Lukas 9, 57-62"
    predigt1_title: str = ""        # "Nicht zurückschauen – ..."
    predigt2_reference: str = ""    # "Epheser 4,23"
    predigt2_title: str = ""        # "Wie das Evangelium ..."
    is_abendmahl: bool = False
    songs: list = field(default_factory=list)  # 7 SongEntry-Objekte
    announcements: list = field(default_factory=list)  # Abkündigungen
    invitation_events: list = field(default_factory=list)  # List[InvitationEvent]


def _get_cell_color(cell) -> Optional[str]:
    """Gibt die RGB-Hintergrundfarbe einer Zelle zurück, oder None."""
    try:
        fill = cell.fill
        if not fill:
            return None
        fg = fill.fgColor
        if fg and fg.type == 'rgb':
            rgb = fg.rgb
            # Transparent/Weiß/Schwarz ignorieren
            if rgb in ("00000000", "FFFFFFFF", "FF000000", "00FFFFFF"):
                return None
            return rgb
    except Exception:
        pass
    return None


def _looks_like_song(val: str) -> bool:
    """Prüft ob ein Zellinhalt wie ein Lied aussieht."""
    if not val:
        return False
    low = val.lower()
    if any(low.startswith(p) for p in _SONG_PREFIXES):
        return True
    if " - " in val and _SONG_BOOKS.match(val):
        return True
    return False


def _looks_like_single_bible_ref(val: str) -> bool:
    """Prüft ob ein Zellinhalt eine einzelne Bibelstelle ist (kein Semikolon)."""
    if not val or ";" in val:
        return False
    return bool(_BIBLE_REF.match(val.strip()))


def _scan_by_color(ws) -> tuple:
    """Scannt das Arbeitsblatt nach Farben und extrahiert Lieder und Bibelstellen.

    Returns:
        (song_raws, lesung_ref, predigt_refs)
        - song_raws: Liste von (row, raw_text, col_b) für alle grünen Lied-Zellen
        - lesung_ref: Erster orangefarbener Bibelstellen-Eintrag
        - predigt_refs: Liste roter Bibelstellen-Einträge (erste=Predigt1, zweite=Predigt2)
    """
    song_raws = []
    lesung_ref = ""
    predigt_refs = []

    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=4)
        color = _get_cell_color(cell)
        val = str(cell.value).strip() if cell.value else ""

        if color == SONG_COLOR and _looks_like_song(val):
            col_b = str(ws.cell(row=row, column=2).value or "").strip()
            song_raws.append((row, val, col_b))
            log.debug(f"Farb-Scan: Lied in Zeile {row}: {val[:60]} (B={col_b})")

        elif color == LESUNG_COLOR and _looks_like_single_bible_ref(val) and not lesung_ref:
            lesung_ref = val
            log.debug(f"Farb-Scan: Lesung in Zeile {row}: {val}")

        elif color == PREDIGT_COLOR and _looks_like_single_bible_ref(val):
            predigt_refs.append(val)
            log.debug(f"Farb-Scan: Predigt-Ref in Zeile {row}: {val}")

    return song_raws, lesung_ref, predigt_refs


def parse_song_entry(raw: str, slot_key: str = "", col_b: str = "") -> SongEntry:
    """Parst einen Lied-Eintrag aus der Excel-Zelle.

    Args:
        raw: Zellinhalt aus Spalte D
        slot_key: Template-Slot (z.B. "song1") – wird ggf. später überschrieben
        col_b: Zellinhalt aus Spalte B für zusätzliche Kategorie-Erkennung
    """
    if not raw or not raw.strip():
        return SongEntry(raw="", slot_key=slot_key)

    song = SongEntry(raw=raw.strip(), slot_key=slot_key)

    text = raw.strip()
    # Doppelte Leerzeichen normalisieren (v1.4.1 Bugfix)
    text = re.sub(r'\s+', ' ', text)

    # Kategorie-Prefix erkennen und entfernen
    if text.lower().startswith("lobpreisstrophe:"):
        song.category = "Lobpreisstrophe"
        text = text[len("lobpreisstrophe:"):].strip()
    elif text.lower().startswith("kinderlied:"):
        song.category = "Kinderlied"
        text = text[len("kinderlied:"):].strip()
    elif text.lower().startswith("sonstige lieder"):
        song.category = "Sonstige Lieder"
        # "Sonstige Lieder - Heilig heilig" oder "Sonstige Lieder: Titel"
        if " - " in text:
            text = text.split(" - ", 1)[1].strip()
        elif ":" in text:
            text = text.split(":", 1)[1].strip()
    else:
        song.category = "Gemeindelied"

    # Spalte B als zusätzliches Signal für Lobpreisstrophe
    # z.B. B="Lobpreisstrophe", D="FJ1 28 - In dir ist mein Leben"
    if song.category == "Gemeindelied" and col_b.lower().startswith("lobpreisstrophe"):
        song.category = "Lobpreisstrophe"

    # Buch und Nummer extrahieren: "FJ1 235 - Jesus, dir nach"
    # Oder nur Titel bei Kinderliedern: "Gottes große Liebe"
    if " - " in text:
        prefix, title = text.split(" - ", 1)
        song.title = title.strip()

        # Prefix parsen: "FJ1 235" oder "SUG 61" oder "FJ 29"
        parts = prefix.strip().split()
        if len(parts) >= 2:
            song.book = parts[0].strip()
            # Nummer: führendes "-" entfernen (v1.4.3 Bugfix)
            num = parts[-1].strip().lstrip("-").rstrip("-")
            # "FJ" ohne Nummer ist manchmal "FJ 29" → book=FJ, number=29
            # Prüfen ob der letzte Teil eine Nummer ist
            if num.replace(".", "").isdigit():
                song.number = num
                # Wenn mehr als 2 Teile, book könnte "FJ1" oder "SGIDH 2" sein
                if len(parts) > 2:
                    song.book = " ".join(parts[:-1])
            else:
                song.book = prefix.strip()
        elif len(parts) == 1:
            # Nur Buch, keine Nummer
            song.book = parts[0].strip()
    else:
        # Kein " - " → reiner Titel (Kinderlieder, etc.)
        song.title = text.strip()

    # Erste 3 Wörter des Titels für die Suche (v1.4.2)
    if song.title:
        words = song.title.split()
        song.title_words = words[:3]

    log.debug(f"Parsed song: {song}")
    return song


def _assign_songs_to_slots(song_raws: list) -> list:
    """Weist Lieder nach liturgischer Rolle den Template-Slots zu.

    Statt Lieder sequentiell song1–song7 zuzuweisen, werden sie nach
    Kategorie auf die richtigen Template-Positionen verteilt:
      - song3: Kinderlied (vor Kinderstunde)
      - song2, song6: Lobpreisstrophen
      - song1, song4, song5, song7: Gemeindelieder (song7 = Schlusslied)

    Args:
        song_raws: Liste von (row, raw_text, col_b) aus dem Farb-Scan
    """
    # Alle Lieder parsen (ohne feste Slot-Zuweisung)
    all_songs = []
    for row, raw, col_b in song_raws:
        song = parse_song_entry(raw, col_b=col_b)
        all_songs.append(song)

    # Nach Kategorie aufteilen
    kinderlieder = [s for s in all_songs if s.category == "Kinderlied"]
    lobpreis = [s for s in all_songs if s.category == "Lobpreisstrophe"]
    gemeinde = [s for s in all_songs
                if s.category in ("Gemeindelied", "Sonstige Lieder")]

    slots = {}

    # Kinderlied → song3
    if kinderlieder:
        slots["song3"] = kinderlieder[0]
        if len(kinderlieder) > 1:
            log.warning(f"{len(kinderlieder)} Kinderlieder gefunden, "
                        f"nur das erste wird verwendet")

    # Lobpreisstrophen → song2, song6
    if len(lobpreis) >= 1:
        slots["song2"] = lobpreis[0]
    if len(lobpreis) >= 2:
        slots["song6"] = lobpreis[1]
    if len(lobpreis) > 2:
        log.warning(f"{len(lobpreis)} Lobpreisstrophen gefunden, "
                    f"nur die ersten 2 werden verwendet")

    # Gemeindelieder → song1 (Eröffnung), song4, song5, song7 (Schlusslied)
    # Überzählige → song_extra1, song_extra2, ...
    if len(gemeinde) >= 2:
        # Erstes Gemeindelied → song1, letztes → song7 (Schlusslied)
        slots["song1"] = gemeinde[0]
        slots["song7"] = gemeinde[-1]
        # Mittlere Gemeindelieder → song4, song5, dann Extras
        middle = gemeinde[1:-1]
        middle_slots = ["song4", "song5"]
        for i, song in enumerate(middle):
            if i < len(middle_slots):
                slots[middle_slots[i]] = song
            else:
                extra_key = f"song_extra{i - len(middle_slots) + 1}"
                slots[extra_key] = song
                log.info(f"Extra-Lied: {song.raw[:40]} → {extra_key}")
    elif len(gemeinde) == 1:
        slots["song1"] = gemeinde[0]
    else:
        log.warning("Keine Gemeindelieder im Farb-Scan gefunden")

    # Ergebnis-Liste aufbauen
    all_slots = ["song1", "song2", "song3", "song4", "song5", "song6", "song7"]
    # Extra-Slots hinzufügen
    extra_slots = sorted([k for k in slots if k.startswith("song_extra")])
    all_slots.extend(extra_slots)

    result = []
    for slot_key in all_slots:
        if slot_key in slots:
            song = slots[slot_key]
            song.slot_key = slot_key
            result.append(song)
        else:
            result.append(SongEntry(raw="", slot_key=slot_key))
            log.warning(f"Slot {slot_key}: kein passendes Lied gefunden")

    # Zusammenfassung loggen
    assigned = [s for s in all_songs if any(s is slots.get(k) for k in all_slots)]
    skipped = [s for s in all_songs if s not in assigned]
    if skipped:
        log.warning(f"{len(skipped)} Lied(er) nicht zugewiesen: "
                    f"{[s.raw[:40] for s in skipped]}")

    return result


def _parse_header(header: str) -> tuple:
    """Extrahiert Datum und Kirchenkalendername aus dem Header.
    'Gottesdienst am 08.03.2026 (Okuli)' → ('08.03.2026', 'Okuli')
    """
    date_str = ""
    kirche = ""

    if not header:
        return date_str, kirche

    # Datum extrahieren: DD.MM.YYYY
    m = re.search(r'(\d{2}\.\d{2}\.\d{4})', header)
    if m:
        date_str = m.group(1)

    # Kirchenkalender extrahieren: (Name)
    m = re.search(r'\(([^)]+)\)', header)
    if m:
        kirche = m.group(1).strip()

    return date_str, kirche


def _check_dropbox_running() -> bool:
    """Prüft ob der Dropbox-Prozess läuft."""
    try:
        result = subprocess.run(["pgrep", "-x", "Dropbox"],
                                capture_output=True, timeout=5)
        return result.returncode == 0
    except Exception:
        return False


def _trigger_dropbox_sync() -> None:
    """Startet Dropbox falls nötig und wartet auf Synchronisation."""
    if not _check_dropbox_running():
        log.info("Dropbox läuft nicht – wird gestartet...")
        subprocess.run(["open", "-a", "Dropbox"], timeout=10)
        # Warten bis Dropbox-Prozess läuft
        for _ in range(15):
            time.sleep(1)
            if _check_dropbox_running():
                log.info("Dropbox gestartet.")
                break
        else:
            raise RuntimeError(
                "Dropbox konnte nicht gestartet werden. "
                "Bitte manuell starten und erneut versuchen."
            )

    # Dropbox-Ordner anfassen um Sync anzustoßen
    log.info("Warte auf Dropbox-Synchronisation...")
    os.listdir(GODI_PLAN_DIR)
    time.sleep(120)
    log.info("Dropbox-Sync abgeschlossen (2 Min. gewartet).")


def _ensure_dropbox_sync(excel_path: str) -> None:
    """Stellt sicher, dass die Excel-Datei aus Dropbox aktuell ist."""
    _trigger_dropbox_sync()

    if not _check_dropbox_running():
        raise RuntimeError(
            "Dropbox läuft NICHT! Die Excel-Datei könnte veraltet sein. "
            "Bitte Dropbox starten und erneut versuchen."
        )

    if GODI_PLAN_DIR not in excel_path:
        raise RuntimeError(
            f"Excel-Datei stammt nicht aus Dropbox ({excel_path}). "
            f"Erwartet wird eine Datei in {GODI_PLAN_DIR}."
        )

    mtime = os.path.getmtime(excel_path)
    mtime_dt = datetime.fromtimestamp(mtime)
    age_hours = (datetime.now() - mtime_dt).total_seconds() / 3600
    log.info(f"Dropbox-Excel zuletzt geändert: {mtime_dt:%d.%m.%Y %H:%M} "
             f"(vor {age_hours:.1f} Stunden)")

    log.info("Dropbox läuft – Excel-Datei ist aktuell.")


def find_godi_plan_excel(sunday: date) -> Optional[str]:
    """Findet die richtige GoDi-Plan Excel-Datei im Dropbox-Ordner."""
    sheet_name = f"So {sunday.day}.{sunday.month}"

    # NUR aus Dropbox lesen – kein Desktop-Fallback
    search_dirs = [GODI_PLAN_DIR]
    for search_dir in search_dirs:
        pattern = os.path.join(search_dir, "GoDi-Plan*.xlsx")
        for path in glob.glob(pattern):
            if path.endswith("~$GoDi-Plan"):  # Lock-Dateien ignorieren
                continue
            if "~$" in os.path.basename(path):
                continue
            try:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                # Prüfe ob Sheet existiert (auch mit trailing space)
                for name in wb.sheetnames:
                    if name.strip() == sheet_name:
                        wb.close()
                        log.info(f"GoDi-Plan gefunden: {path}, Sheet: {name}")
                        return path
                wb.close()
            except Exception as e:
                log.warning(f"Fehler beim Öffnen von {path}: {e}")

    return None


def list_all_sheets() -> list[tuple[str, str]]:
    """Listet alle Sheets aus allen GoDi-Plan Excel-Dateien.

    Returns:
        Liste von (sheet_name, excel_path) Tupeln.
        Hilfs-Sheets (Überblick, Vorlage, etc.) werden gefiltert.
    """
    skip = {"Überblick", "GoDi-Vorlage", "Nächsten GoDis im nächsten Plan"}
    results = []
    pattern = os.path.join(GODI_PLAN_DIR, "GoDi-Plan*.xlsx")
    for path in glob.glob(pattern):
        if "~$" in os.path.basename(path):
            continue
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for name in wb.sheetnames:
                stripped = name.strip()
                if stripped and stripped not in skip:
                    results.append((stripped, path))
            wb.close()
        except Exception as e:
            log.warning(f"Fehler beim Öffnen von {path}: {e}")
    return results


def read_godi_plan_by_sheet(sheet_name: str, excel_path: str,
                             skip_dropbox_sync: bool = False) -> GodiPlanData | None:
    """Liest den GoDi-Plan für ein beliebiges Sheet.

    Args:
        sheet_name: Name des Sheets (z.B. 'So 15.3', 'Pa 18.3')
        excel_path: Pfad zur Excel-Datei
        skip_dropbox_sync: True für interaktive Nutzung (kein 120s Wait)
    """
    if not skip_dropbox_sync:
        _ensure_dropbox_sync(excel_path)

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Sheet finden (mit trailing space Fallback)
    ws = None
    for name in wb.sheetnames:
        if name.strip() == sheet_name:
            ws = wb[name]
            break

    if ws is None:
        log.error(f"Sheet '{sheet_name}' nicht gefunden in {excel_path}")
        wb.close()
        return None

    def cell(key: str) -> str:
        """Liest eine Zelle anhand des Schlüssels aus EXCEL_ROWS."""
        row, col = EXCEL_ROWS[key]
        val = ws.cell(row=row, column=col).value
        return str(val).strip() if val else ""

    data = GodiPlanData()

    # Header und Metadaten (feste Positionen, da diese sich selten verschieben)
    data.service_header = cell("header")
    data.theme = cell("theme")
    data.date_str, data.kirchenkalender = _parse_header(data.service_header)

    # Begrüßungsvers
    data.greeting_verse = cell("begruessung")

    # === Farb-basierter Scan für Lieder und Bibelstellen ===
    song_raws, lesung_ref_color, predigt_refs_color = _scan_by_color(ws)

    # Bibelstellen: Farb-Scan bevorzugt, Fallback auf feste Zeilen
    if lesung_ref_color:
        data.lesung_reference = lesung_ref_color
        log.info(f"Lesung (Farbe erkannt): {lesung_ref_color}")
    else:
        data.lesung_reference = cell("lesung_referenz")
        log.warning("Lesung: Farb-Erkennung fehlgeschlagen, nutze feste Zeile")

    if len(predigt_refs_color) >= 1:
        data.predigt1_reference = predigt_refs_color[0]
        log.info(f"Predigt1 (Farbe erkannt): {predigt_refs_color[0]}")
    else:
        data.predigt1_reference = cell("predigt1_referenz")
        log.warning("Predigt1: Farb-Erkennung fehlgeschlagen, nutze feste Zeile")

    if len(predigt_refs_color) >= 2:
        data.predigt2_reference = predigt_refs_color[1]
        log.info(f"Predigt2 (Farbe erkannt): {predigt_refs_color[1]}")
    else:
        data.predigt2_reference = cell("predigt2_referenz")
        if len(predigt_refs_color) < 2:
            log.warning("Predigt2: Farb-Erkennung fehlgeschlagen, nutze feste Zeile")

    # Predigt-Titel und -Referenz dynamisch über Spalte B suchen
    # (feste Zeilen funktionieren nicht bei variablen Gottesdienst-Layouts)
    predigt1_found = False
    predigt2_found = False
    for row in range(1, ws.max_row + 1):
        col_b = ws.cell(row=row, column=2).value
        if not col_b or not isinstance(col_b, str):
            continue
        col_b_lower = col_b.strip().lower()

        if "predigt 1" in col_b_lower or "predigt1" in col_b_lower:
            ref = ws.cell(row=row, column=4).value
            title = ws.cell(row=row, column=5).value
            if ref:
                data.predigt1_reference = str(ref).strip()
            if title:
                data.predigt1_title = str(title).strip()
            predigt1_found = True
            log.info(f"Predigt1 (Spalte B Zeile {row}): ref=\"{data.predigt1_reference}\" titel=\"{data.predigt1_title}\"")

        elif "predigt 2" in col_b_lower or "predigt2" in col_b_lower:
            ref = ws.cell(row=row, column=4).value
            title = ws.cell(row=row, column=5).value
            if ref:
                data.predigt2_reference = str(ref).strip()
            if title:
                data.predigt2_title = str(title).strip()
            predigt2_found = True
            log.info(f"Predigt2 (Spalte B Zeile {row}): ref=\"{data.predigt2_reference}\" titel=\"{data.predigt2_title}\"")

    if not predigt1_found:
        # Fallback auf feste Zeile
        if not data.predigt1_title:
            data.predigt1_title = cell("predigt1_titel")
    if not predigt2_found:
        if not data.predigt2_title:
            data.predigt2_title = cell("predigt2_titel")

    # Abendmahl-Erkennung
    abendmahl_val = cell("abendmahl")
    data.is_abendmahl = "abendmahl" in abendmahl_val.lower() if abendmahl_val else False

    # === Lieder: Farb-Scan ===
    if len(song_raws) >= 2:
        log.info(f"Farb-Scan: {len(song_raws)} Lieder gefunden")
        data.songs = _assign_songs_to_slots(song_raws)
    else:
        # Fallback: feste Zeilen-Mappings
        log.warning(f"Farb-Scan lieferte nur {len(song_raws)} Lieder – "
                    f"Fallback auf feste Zeilen-Mappings")
        song_keys = ["song1", "song2", "song3", "song4", "song5", "song6", "song7"]
        for key in song_keys:
            raw = cell(key)
            song = parse_song_entry(raw, key)
            data.songs.append(song)

    # Abkündigungen
    for row_num in ABKUENDIGUNGEN_ROWS:
        parts = []
        for col in range(2, 6):  # Cols B-E
            val = ws.cell(row=row_num, column=col).value
            if val:
                parts.append(str(val).strip())
        if parts:
            data.announcements.append(" | ".join(parts))

    # Einladungs-Events (Herzliche Einladung Folie)
    for row_num in range(EINLADUNG_ROW_START, EINLADUNG_ROW_END + 1):
        datum_val = ws.cell(row=row_num, column=EINLADUNG_COLS["datum"]).value
        if not datum_val:
            continue
        zeit_val = ws.cell(row=row_num, column=EINLADUNG_COLS["uhrzeit"]).value
        event_val = ws.cell(row=row_num, column=EINLADUNG_COLS["veranstaltung"]).value
        hinweis_val = ws.cell(row=row_num, column=EINLADUNG_COLS["hinweis"]).value

        # Datum: String beibehalten ("Di 17.03.26")
        datum_str = str(datum_val).strip() if datum_val else ""

        # Uhrzeit: datetime.time → "HH:MM"
        if hasattr(zeit_val, 'strftime'):
            zeit_str = zeit_val.strftime("%H:%M")
        elif zeit_val:
            zeit_str = str(zeit_val).strip()
            # "19:00:00" → "19:00"
            if zeit_str.count(":") == 2:
                zeit_str = zeit_str[:5]
        else:
            zeit_str = ""

        event_str = str(event_val).strip() if event_val else ""
        hinweis_str = str(hinweis_val).strip() if hinweis_val else ""

        if datum_str and event_str:
            data.invitation_events.append(InvitationEvent(
                date_str=datum_str,
                time_str=zeit_str,
                event_name=event_str,
                note=hinweis_str,
            ))
            log.debug(f"Einladung: {datum_str} {zeit_str} {event_str}"
                      f"{' (' + hinweis_str + ')' if hinweis_str else ''}")

    log.info(f"Einladungs-Events: {len(data.invitation_events)} gefunden")

    wb.close()

    log.info(f"GoDi-Plan gelesen: {data.service_header}")
    log.info(f"Thema: {data.theme}")
    log.info(f"Lieder: {[s.raw for s in data.songs]}")

    return data


def read_godi_plan(sunday: date) -> GodiPlanData | None:
    """Liest den GoDi-Plan für den gegebenen Sonntag (Wrapper für Automation)."""
    excel_path = find_godi_plan_excel(sunday)
    if not excel_path:
        log.error(f"Keine GoDi-Plan Datei gefunden für {sunday}")
        return None
    sheet_name = f"So {sunday.day}.{sunday.month}"
    return read_godi_plan_by_sheet(sheet_name, excel_path)


if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG)
    # Test mit So 15.3.2026
    result = read_godi_plan(date(2026, 3, 15))
    if result:
        print(f"\nHeader: {result.service_header}")
        print(f"Thema: {result.theme}")
        print(f"Datum: {result.date_str}, Kirche: {result.kirchenkalender}")
        print(f"Lesung: {result.lesung_reference}")
        print(f"Predigt 1: {result.predigt1_reference} - {result.predigt1_title}")
        print(f"Predigt 2: {result.predigt2_reference} - {result.predigt2_title}")
        print(f"Abendmahl: {result.is_abendmahl}")
        print(f"\nLieder:")
        for s in result.songs:
            print(f"  {s.slot_key}: book={s.book}, num={s.number}, "
                  f"title={s.title}, cat={s.category}")
