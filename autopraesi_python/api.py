"""AutoPräsi API – FastAPI Backend für die Web-UI."""
from __future__ import annotations

import io
import logging
import os
from dataclasses import asdict
from typing import Optional
from urllib.parse import quote

import openpyxl
from fastapi import FastAPI, UploadFile, File, HTTPException, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

import storage
from config import (IMAGE_DIR, OUTPUT_DIR_DROPBOX, GODI_PLAN_DIR,
                    TOGGLEABLE_SECTIONS, DEFAULT_SECTION_ORDER)
from excel_reader import (list_all_sheets, read_godi_plan_by_sheet, GodiPlanData,
                          parse_song_entry, find_godi_plan_excel, _is_godi_plan)
from song_finder import build_song_index, find_song
from presentation_builder import build_presentation
import godi_editor

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
log = logging.getLogger("autopraesi.api")

app = FastAPI(title="AutoPräsi API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Song-Index einmalig bauen und cachen
_song_index = None

# Datei-Cache für den GoDi-Plan-Editor: vermeidet wiederholte Dropbox-Downloads
# beim Mappenwechsel. Schlüssel = Pfad, Wert = (rev, bytes). Über die Dropbox-rev
# wird der Cache automatisch ungültig, sobald sich die Datei ändert.
_godi_file_cache: dict[str, tuple] = {}


def _godi_get_bytes(path: str) -> tuple[Optional[str], bytes]:
    """Liefert (rev, bytes) der Datei – aus dem Cache, falls die rev unverändert ist.

    Die rev-Abfrage ist ein kleiner Metadaten-Aufruf; der eigentliche (große)
    Download passiert nur beim ersten Mal bzw. nach einer Änderung.
    """
    rev = storage.get_rev(path)
    ent = _godi_file_cache.get(path)
    if ent and rev is not None and ent[0] == rev:
        return rev, ent[1]
    data = storage.download_bytes(path)
    _godi_file_cache[path] = (rev, data)
    return rev, data


def _get_song_index():
    global _song_index
    if _song_index is None:
        _song_index = build_song_index()
    return _song_index


def _find_image(date_str: str) -> Optional[str]:
    """Sucht das Hintergrundbild ('Bild 22.03.jpg' oder '22.3.jpg')."""
    if not date_str:
        return None
    parts = date_str.split(".")
    if len(parts) < 3:
        return None
    day, month = parts[0], parts[1]
    day_short = str(int(day))
    month_short = str(int(month))
    for name in [f"Bild {day}.{month}.jpg", f"{day_short}.{month_short}.jpg"]:
        path = f"{IMAGE_DIR}/{name}"
        if storage.file_exists(path):
            return path
    return None


# --- Models ---

class SheetInfo(BaseModel):
    name: str
    excel_path: str


class SongStatus(BaseModel):
    slot_key: str
    raw: str
    category: str
    book: str
    number: str
    title: str
    found: bool
    file_name: str


class SectionInfo(BaseModel):
    key: str
    label: str
    default_enabled: bool


class GenerateRequest(BaseModel):
    sheet_name: str
    excel_path: str
    overrides: Optional[dict] = None
    fetch_bible: bool = True
    disabled_sections: Optional[list[str]] = None  # z.B. ["glaubensbekenntnis", "kinderstunde"]
    section_order: Optional[list[str]] = None  # z.B. ["begruessung", "song1", ...]
    text_color: str = "white"  # "white" oder "black"
    title_layout: Optional[dict] = None  # {x, y, w, h, fontSize} in %
    subtitle_layout: Optional[dict] = None
    image_filter: str = "none"  # "none", "dark-30", "dark-50", "dark-70", "light-30", "light-50", "gradient-bottom", "gradient-top"
    text_banner: str = "none"  # "none", "subtle", "medium", "strong"
    shadow_strength: str = "normal"  # "normal", "strong"
    text_outline: bool = False


# --- Endpoints ---

def _current_quarter_pattern() -> str:
    """Gibt das Dateinamen-Muster für das aktuelle Quartal zurück."""
    import datetime
    now = datetime.date.today()
    q = (now.month - 1) // 3 + 1
    year = now.year
    # Q1 kann als "_1" oder "_Q1" benannt sein
    if q == 1:
        return f"{year}_1"
    return f"{year}_Q{q}"


@app.get("/api/sheets", response_model=list[SheetInfo])
def get_sheets():
    """Alle verfügbaren Sheets aus den GoDi-Plan Excel-Dateien."""
    sheets = list_all_sheets()
    return [SheetInfo(name=name, excel_path=path) for name, path in sheets]


@app.get("/api/current-quarter")
def get_current_quarter():
    """Gibt das Dateinamen-Muster des aktuellen Quartals zurück."""
    return {"pattern": _current_quarter_pattern()}


@app.get("/api/sections", response_model=list[SectionInfo])
def get_sections():
    """Alle Abschnitte in der Standard-Reihenfolge (für Drag & Drop)."""
    result = []
    for key in DEFAULT_SECTION_ORDER:
        if key in TOGGLEABLE_SECTIONS:
            result.append(SectionInfo(
                key=key, label=TOGGLEABLE_SECTIONS[key]["label"],
                default_enabled=True,
            ))
    return result


@app.get("/api/sheet/{sheet_name}")
def get_sheet_data(sheet_name: str, excel_path: str):
    """Liest die Daten eines Sheets und gibt sie als JSON zurück."""
    data = read_godi_plan_by_sheet(sheet_name, excel_path)
    if not data:
        raise HTTPException(404, f"Sheet '{sheet_name}' nicht gefunden")

    index = _get_song_index()
    songs = []
    for song in data.songs:
        path = find_song(song, index)
        songs.append(SongStatus(
            slot_key=song.slot_key,
            raw=song.raw,
            category=song.category,
            book=song.book,
            number=song.number,
            title=song.title,
            found=path is not None,
            file_name=os.path.basename(path) if path else "",
        ))

    image_path = _find_image(data.date_str)

    return {
        "service_header": data.service_header,
        "theme": data.theme,
        "date_str": data.date_str,
        "kirchenkalender": data.kirchenkalender,
        "greeting_verse": data.greeting_verse,
        "lesung_reference": data.lesung_reference,
        "predigt1_reference": data.predigt1_reference,
        "predigt1_title": data.predigt1_title,
        "predigt2_reference": data.predigt2_reference,
        "predigt2_title": data.predigt2_title,
        "is_abendmahl": data.is_abendmahl,
        "songs": [s.model_dump() for s in songs],
        "announcements": data.announcements,
        "invitation_events": [asdict(e) for e in data.invitation_events],
        "image_found": image_path is not None,
        "image_path": image_path,
    }


@app.get("/api/sheet/{sheet_name}/rows")
def get_sheet_rows(sheet_name: str, excel_path: str):
    """Gibt die rohen Excel-Zeilen eines Sheets zurück (Uhrzeit, Programm, Details)."""
    import datetime
    wb = openpyxl.load_workbook(io.BytesIO(storage.download_bytes(excel_path)), data_only=True)
    ws = None
    for name in wb.sheetnames:
        if name.strip() == sheet_name:
            ws = wb[name]
            break
    if ws is None:
        wb.close()
        raise HTTPException(404, f"Sheet '{sheet_name}' nicht gefunden")

    def fmt(val):
        if val is None:
            return ""
        if isinstance(val, (datetime.time, datetime.datetime)):
            return val.strftime("%H:%M")
        return str(val).strip()

    rows = []
    for r in range(1, ws.max_row + 1):
        uhrzeit = fmt(ws.cell(row=r, column=1).value)
        programm = fmt(ws.cell(row=r, column=2).value)
        details = fmt(ws.cell(row=r, column=4).value)
        if not uhrzeit and not programm and not details:
            continue
        rows.append({
            "row": r,
            "uhrzeit": uhrzeit,
            "programmpunkt": programm,
            "details": details,
        })
    wb.close()
    return rows


@app.get("/api/search-song")
def search_song(raw: str):
    """Sucht ein Lied in der Bibliothek anhand des Rohtexts."""
    if not raw.strip():
        return {"found": False, "file_name": "", "path": ""}
    song = parse_song_entry(raw.strip())
    index = _get_song_index()
    path = find_song(song, index)
    return {
        "found": path is not None,
        "file_name": os.path.basename(path) if path else "",
        "path": path or "",
    }


@app.get("/api/image")
def get_image(path: str):
    """Gibt ein Bild aus Dropbox zurück (für Vorschau im Frontend)."""
    if not storage.file_exists(path):
        raise HTTPException(404, "Bild nicht gefunden")
    data = storage.download_bytes(path)
    media = "image/png" if path.lower().endswith(".png") else "image/jpeg"
    return Response(content=data, media_type=media)


@app.post("/api/upload-image")
async def upload_image(file: UploadFile = File(...)):
    """Lädt ein Bild nach Dropbox hoch und gibt den Dropbox-Pfad zurück.

    Ablage unter /Gemeinde/_uploads, damit das Bild den Request-Übergang
    (Upload → Generate, ggf. anderer Cloud-Worker) überlebt.
    """
    content = await file.read()
    filename = file.filename or "upload.jpg"
    dest = f"{IMAGE_DIR}/_uploads/{filename}"
    storage.upload_bytes(content, dest)
    return {"path": dest, "filename": filename}


@app.post("/api/generate")
def generate_presentation(req: GenerateRequest):
    """Generiert die Präsentation."""
    data = read_godi_plan_by_sheet(req.sheet_name, req.excel_path)
    if not data:
        raise HTTPException(404, f"Sheet '{req.sheet_name}' nicht gefunden")

    # Overrides anwenden
    if req.overrides:
        o = req.overrides
        if "theme" in o and o["theme"]:
            data.theme = o["theme"]
        if "greeting_verse" in o and o["greeting_verse"]:
            data.greeting_verse = o["greeting_verse"]
        if "lesung_reference" in o and o["lesung_reference"]:
            data.lesung_reference = o["lesung_reference"]
        if "predigt1_reference" in o and o["predigt1_reference"]:
            data.predigt1_reference = o["predigt1_reference"]
        if "predigt1_title" in o and o["predigt1_title"]:
            data.predigt1_title = o["predigt1_title"]
        if "predigt2_reference" in o and o["predigt2_reference"]:
            data.predigt2_reference = o["predigt2_reference"]
        if "predigt2_title" in o and o["predigt2_title"]:
            data.predigt2_title = o["predigt2_title"]
        if "announcements" in o:
            data.announcements = o["announcements"]
        if "invitation_events" in o:
            from excel_reader import InvitationEvent
            data.invitation_events = [
                InvitationEvent(
                    date_str=e.get("date_str", ""),
                    time_str=e.get("time_str", ""),
                    event_name=e.get("event_name", ""),
                    note=e.get("note", ""),
                )
                for e in o["invitation_events"]
            ]

        # Song-Overrides (inkl. neue Extra-Songs)
        if "songs" in o and o["songs"]:
            for slot_key, raw_text in o["songs"].items():
                found = False
                for i, song in enumerate(data.songs):
                    if song.slot_key == slot_key:
                        data.songs[i] = parse_song_entry(raw_text, slot_key)
                        found = True
                        break
                # Neuer Extra-Song (manuell hinzugefügt)
                if not found and slot_key.startswith("song_extra") and raw_text.strip():
                    data.songs.append(parse_song_entry(raw_text, slot_key))

    # Skip-Slides berechnen aus disabled_sections
    skip_slides = set()
    if req.disabled_sections:
        for section_key in req.disabled_sections:
            if section_key in TOGGLEABLE_SECTIONS:
                skip_slides.update(TOGGLEABLE_SECTIONS[section_key]["slides"])
                log.info(f"Abschnitt deaktiviert: {section_key} "
                         f"(Folien {TOGGLEABLE_SECTIONS[section_key]['slides']})")

    # Songs suchen
    index = _get_song_index()
    song_paths = {}
    missing = []
    for song in data.songs:
        path = find_song(song, index)
        if path:
            song_paths[song.slot_key] = path
        elif song.raw:
            missing.append(song.slot_key)

    # Bild
    image_path = req.overrides.get("image_path") if req.overrides else None
    if not image_path:
        image_path = _find_image(data.date_str)

    # Output-Name aus Sheet-Name ableiten
    output_name = f"{req.sheet_name}_ungeprüft.pptx"

    try:
        output = build_presentation(
            data, song_paths,
            image_path=image_path,
            fetch_bible=req.fetch_bible,
            output_name=output_name,
            skip_slides=skip_slides if skip_slides else None,
            section_order=req.section_order,
            text_color=req.text_color,
            title_layout=req.title_layout,
            subtitle_layout=req.subtitle_layout,
            image_filter=req.image_filter,
            text_banner=req.text_banner,
            shadow_strength=req.shadow_strength,
            text_outline=req.text_outline,
        )
    except Exception as e:
        log.error(f"Fehler beim Generieren: {e}", exc_info=True)
        raise HTTPException(500, f"Fehler beim Generieren: {e}")

    return {
        "success": True,
        "output_path": output,
        "output_name": os.path.basename(output),
        "missing_songs": missing,
    }


@app.get("/api/download/{filename}")
def download_file(filename: str):
    """Generierte Präsentation aus Dropbox herunterladen."""
    dbx_path = f"{OUTPUT_DIR_DROPBOX}/{filename}"
    if not storage.file_exists(dbx_path):
        raise HTTPException(404, f"Datei nicht gefunden: {filename}")
    data = storage.download_bytes(dbx_path)
    # RFC 5987: nicht-ASCII Dateinamen (z.B. "ungeprüft") korrekt kodieren
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@app.post("/api/upload-excel")
async def upload_excel(file: UploadFile = File(...)):
    """Lädt eine Excel-Datei nach Dropbox hoch und gibt den Dropbox-Pfad zurück."""
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Nur .xlsx Dateien erlaubt")
    content = await file.read()
    dest = f"{GODI_PLAN_DIR}/{file.filename}"
    storage.upload_bytes(content, dest)
    log.info(f"Excel hochgeladen: {dest}")
    return {"success": True, "path": dest, "filename": file.filename}


# ===================================================================
# GoDi-Plan Editor – die ganze Excel als editierbares Raster ("Reiter")
# ===================================================================

class GodiSaveRequest(BaseModel):
    excel_path: str
    sheet: str
    operations: list[dict]
    base_rev: Optional[str] = None  # Dropbox-rev beim Laden (Konflikt-Erkennung)


@app.get("/api/godi/files")
def godi_files():
    """Alle GoDi-Plan Excel-Dateien in Dropbox (Name, Pfad, aktuelles-Quartal-Flag)."""
    pattern = _current_quarter_pattern()
    files = []
    for name, path in storage.list_files(GODI_PLAN_DIR, suffix=".xlsx"):
        if not _is_godi_plan(name):
            continue
        files.append({
            "name": name,
            "excel_path": path,
            "is_current_quarter": pattern in path,
        })
    files.sort(key=lambda f: f["name"])
    return files


@app.get("/api/godi/sheets")
def godi_sheets(excel_path: str):
    """Alle Tabellenblätter (Mappen / Sonntage) einer Excel-Datei – in Originalreihenfolge."""
    skip = {"Überblick", "GoDi-Vorlage", "Nächsten GoDis im nächsten Plan"}
    _rev, data = _godi_get_bytes(excel_path)
    names = godi_editor.list_sheets(data)
    return [{"name": n, "is_helper": n in skip} for n in names]


@app.get("/api/godi/upcoming-sunday")
def godi_upcoming_sunday():
    """Ermittelt die Datei + Mappe für den kommenden Sonntag dieser Woche.

    Ist heute Sonntag, zählt heute. Gibt {excel_path, sheet} zurück oder
    {excel_path: null} wenn keine passende Mappe gefunden wurde.
    """
    import datetime
    today = datetime.date.today()
    days_until_sunday = (6 - today.weekday()) % 7  # Montag=0 … Sonntag=6
    sunday = today + datetime.timedelta(days=days_until_sunday)
    sheet_name = f"So {sunday.day}.{sunday.month}"
    excel_path = find_godi_plan_excel(sunday)
    return {
        "excel_path": excel_path,
        "sheet": sheet_name if excel_path else None,
        "date": sunday.isoformat(),
    }


@app.get("/api/godi/grid")
def godi_grid(excel_path: str, sheet: str):
    """Liefert ein komplettes Tabellenblatt als Raster (Werte + Ansicht)."""
    rev, data = _godi_get_bytes(excel_path)
    grid = godi_editor.read_grid(data, sheet)
    if grid is None:
        raise HTTPException(404, f"Blatt '{sheet}' nicht gefunden in {excel_path}")
    grid["rev"] = rev
    return grid


@app.post("/api/godi/save")
def godi_save(req: GodiSaveRequest):
    """Wendet Bearbeitungen an und schreibt die Excel zurück nach Dropbox.

    Schritte: Konfliktprüfung (rev) → Auto-Backup → Operationen anwenden → Upload.
    """
    if not req.operations:
        return {"success": True, "changed": 0, "rev": storage.get_rev(req.excel_path)}

    # Konflikt-Erkennung: wurde die Datei seit dem Laden verändert?
    current_rev = storage.get_rev(req.excel_path)
    if req.base_rev and current_rev and req.base_rev != current_rev:
        raise HTTPException(
            409,
            "Die Datei wurde seit dem Öffnen geändert (z.B. im Dropbox-Desktop). "
            "Bitte den Reiter neu laden, damit deine Änderungen nichts überschreiben.",
        )

    file_bytes = storage.download_bytes(req.excel_path)

    # Auto-Backup der unveränderten Datei (Sicherheitsnetz)
    import datetime
    stamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    base_name = req.excel_path.split("/")[-1].rsplit(".", 1)[0]
    backup_path = f"{GODI_PLAN_DIR}/_backups/{base_name}_{stamp}.xlsx"
    try:
        storage.copy_file(req.excel_path, backup_path)
    except Exception as e:
        log.warning(f"Backup fehlgeschlagen (fahre fort): {e}")
        backup_path = None

    try:
        new_bytes = godi_editor.apply_operations(file_bytes, req.sheet, req.operations)
    except ValueError as e:
        raise HTTPException(404, str(e))
    except Exception as e:
        log.error(f"Fehler beim Anwenden der Operationen: {e}", exc_info=True)
        raise HTTPException(500, f"Bearbeitung fehlgeschlagen: {e}")

    storage.upload_bytes(new_bytes, req.excel_path)
    log.info(f"GoDi-Plan gespeichert: {req.excel_path} ({len(req.operations)} Operationen)")

    # Cache mit den frisch gespeicherten Bytes + neuer rev aktualisieren
    new_rev = storage.get_rev(req.excel_path)
    _godi_file_cache[req.excel_path] = (new_rev, new_bytes)

    return {
        "success": True,
        "changed": len(req.operations),
        "backup": backup_path,
        "rev": new_rev,
    }


@app.post("/api/refresh-songs")
def refresh_song_index():
    """Song-Index neu aufbauen."""
    global _song_index
    _song_index = build_song_index()
    count = sum(len(v) for v in _song_index.values())
    return {"success": True, "total_songs": count}
