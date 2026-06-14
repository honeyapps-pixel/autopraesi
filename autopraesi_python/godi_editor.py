"""Liest und schreibt GoDi-Plan Excel-Blätter als editierbares Raster.

Anders als excel_reader.py (extrahiert nur die für die Präsentation relevanten
Felder) liefert dieses Modul das *gesamte* Tabellenblatt inklusive Ansicht –
Zellwerte, Hintergrundfarben, Schrift, Ausrichtung, verbundene Zellen,
Spaltenbreiten und Zeilenhöhen – für die Excel-ähnliche Web-Tabelle.

Lesen geschieht mit ``data_only=True`` (zeigt berechnete Werte statt Formeln).
Schreiben lädt die Datei *ohne* ``data_only`` und wendet nur die übergebenen
Operationen an; alle übrigen Zellen (inkl. Formeln und Formatierung) bleiben
dadurch unangetastet.
"""
from __future__ import annotations

import datetime
import io
import logging
import re
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

log = logging.getLogger(__name__)

# Excel-Standardbreite (Zeichen) bzw. -höhe (Punkt), wenn nichts gesetzt ist.
_DEFAULT_COL_WIDTH = 8.43
_DEFAULT_ROW_HEIGHT = 15.0

_TIME_RE = re.compile(r"^([01]?\d|2[0-3]):([0-5]\d)(:[0-5]\d)?$")


def _find_sheet(wb, sheet_name: str):
    """Findet ein Blatt anhand des Namens (mit Fallback auf trailing space)."""
    for name in wb.sheetnames:
        if name.strip() == sheet_name.strip():
            return wb[name]
    return None


def list_sheets(file_bytes: bytes) -> list[str]:
    """Gibt alle Tabellenblatt-Namen einer Excel-Datei zurück (Originalreihenfolge)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    names = [n.strip() for n in wb.sheetnames]
    wb.close()
    return names


def _color_hex(color) -> Optional[str]:
    """Wandelt eine openpyxl-Farbe in '#RRGGBB' um (nur explizite RGB-Farben).

    Theme-/Indexfarben werden übersprungen (None), da ihre Umrechnung unzuverlässig
    ist; die GoDi-Pläne nutzen durchgängig explizite RGB-Füllungen.
    """
    if color is None:
        return None
    if getattr(color, "type", None) != "rgb":
        return None
    rgb = color.rgb
    if not isinstance(rgb, str) or len(rgb) != 8:
        return None
    # Transparent ignorieren
    if rgb in ("00000000",):
        return None
    return "#" + rgb[2:].upper()


def _format_value(cell) -> str:
    """Erzeugt die Anzeige-Zeichenkette einer Zelle (wie in Excel sichtbar)."""
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M")
    if isinstance(v, datetime.date):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, float):
        # Ganzzahlige Floats ohne ".0" anzeigen
        if v.is_integer():
            return str(int(v))
        return str(v)
    return str(v)


def read_grid(file_bytes: bytes, sheet_name: str) -> Optional[dict]:
    """Liest ein komplettes Tabellenblatt als JSON-fähiges Raster.

    Returns None, wenn das Blatt nicht existiert.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = _find_sheet(wb, sheet_name)
    if ws is None:
        wb.close()
        return None

    max_row = max(ws.max_row, 1)
    max_col = max(ws.max_column, 1)

    # Spaltenbreiten (Excel-Zeichen-Einheit) → Pixel beim Frontend
    col_widths: dict[int, float] = {}
    for letter, dim in ws.column_dimensions.items():
        if dim.width and dim.customWidth:
            try:
                col_widths[column_index_from_string(letter)] = dim.width
            except ValueError:
                continue

    # Zeilenhöhen (Punkt)
    row_heights: dict[int, float] = {}
    for idx, dim in ws.row_dimensions.items():
        if dim.height:
            row_heights[idx] = dim.height

    merged = [
        {"r1": rng.min_row, "c1": rng.min_col, "r2": rng.max_row, "c2": rng.max_col}
        for rng in ws.merged_cells.ranges
    ]

    cells: dict[str, dict] = {}
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            value = _format_value(cell)
            style: dict = {}

            font = cell.font
            if font:
                if font.bold:
                    style["b"] = True
                if font.italic:
                    style["i"] = True
                if font.size and font.size != 11:
                    style["sz"] = float(font.size)
                fg = _color_hex(font.color)
                if fg and fg != "#000000":
                    style["fg"] = fg

            if cell.fill and cell.fill.patternType == "solid":
                bg = _color_hex(cell.fill.fgColor)
                if bg:
                    style["bg"] = bg

            align = cell.alignment
            if align:
                if align.horizontal:
                    style["a"] = align.horizontal
                if align.vertical and align.vertical != "bottom":
                    style["va"] = align.vertical
                if align.wrap_text:
                    style["wrap"] = True

            border = cell.border
            if border:
                if border.top and border.top.style:
                    style["bt"] = True
                if border.right and border.right.style:
                    style["br"] = True
                if border.bottom and border.bottom.style:
                    style["bb"] = True
                if border.left and border.left.style:
                    style["bl"] = True

            if value == "" and not style:
                continue
            entry: dict = {}
            if value != "":
                entry["v"] = value
            if style:
                entry["s"] = style
            cells[f"{r}:{c}"] = entry

    freeze = ws.freeze_panes  # z.B. "A2" oder None
    wb.close()

    return {
        "sheet": sheet_name,
        "max_row": max_row,
        "max_col": max_col,
        "col_widths": col_widths,
        "row_heights": row_heights,
        "default_col_width": _DEFAULT_COL_WIDTH,
        "default_row_height": _DEFAULT_ROW_HEIGHT,
        "merged": merged,
        "freeze": freeze,
        "cells": cells,
    }


def _coerce(value: str):
    """Wandelt eine vom Frontend gesendete Zeichenkette in einen Excel-Wert um."""
    if value is None:
        return None
    s = str(value)
    if s == "":
        return None
    # Uhrzeit "HH:MM"
    m = _TIME_RE.match(s.strip())
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        sec = int(m.group(3)[1:]) if m.group(3) else 0
        return datetime.time(hour=h, minute=mi, second=sec)
    # Zahl (nur reine Ziffern / Dezimalzahl, kein führendes '0' bei mehrstellig wie PLZ)
    stripped = s.strip()
    if re.fullmatch(r"-?\d+", stripped) and not (len(stripped) > 1 and stripped.lstrip("-").startswith("0")):
        try:
            return int(stripped)
        except ValueError:
            pass
    if re.fullmatch(r"-?\d+[.,]\d+", stripped):
        try:
            return float(stripped.replace(",", "."))
        except ValueError:
            pass
    return s


def apply_operations(file_bytes: bytes, sheet_name: str, operations: list[dict]) -> bytes:
    """Wendet eine Liste von Bearbeitungs-Operationen an und gibt neue Datei-Bytes zurück.

    Unterstützte Operationen (Feld ``op``):
      - ``set``       {row, col, value}                  – Zellwert setzen
      - ``format``    {row, col, bold?, italic?, color?, fill?, align?} – Formatierung
      - ``insertRow`` {index, count?}                    – Zeile(n) einfügen
      - ``deleteRow`` {index, count?}                    – Zeile(n) löschen
      - ``insertCol`` {index, count?}                    – Spalte(n) einfügen
      - ``deleteCol`` {index, count?}                    – Spalte(n) löschen
      - ``merge``     {r1, c1, r2, c2}                   – Zellen verbinden
      - ``unmerge``   {r1, c1, r2, c2}                   – Verbindung lösen
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))  # mit Formeln + Stilen
    ws = _find_sheet(wb, sheet_name)
    if ws is None:
        raise ValueError(f"Blatt '{sheet_name}' nicht gefunden")

    for op in operations:
        kind = op.get("op")
        if kind == "set":
            ws.cell(row=op["row"], column=op["col"]).value = _coerce(op.get("value"))

        elif kind == "format":
            cell = ws.cell(row=op["row"], column=op["col"])
            f = cell.font
            cell.font = Font(
                name=f.name, size=f.size,
                bold=op.get("bold", f.bold),
                italic=op.get("italic", f.italic),
                color=op["color"].replace("#", "FF") if op.get("color") else f.color,
            )
            if "fill" in op:
                if op["fill"]:
                    rgb = "FF" + op["fill"].replace("#", "")
                    cell.fill = PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")
                else:
                    cell.fill = PatternFill(fill_type=None)
            if "align" in op:
                a = cell.alignment
                cell.alignment = Alignment(
                    horizontal=op["align"] or None,
                    vertical=a.vertical, wrap_text=a.wrap_text,
                )

        elif kind == "insertRow":
            ws.insert_rows(op["index"], op.get("count", 1))
        elif kind == "deleteRow":
            ws.delete_rows(op["index"], op.get("count", 1))
        elif kind == "insertCol":
            ws.insert_cols(op["index"], op.get("count", 1))
        elif kind == "deleteCol":
            ws.delete_cols(op["index"], op.get("count", 1))

        elif kind == "merge":
            ws.merge_cells(start_row=op["r1"], start_column=op["c1"],
                           end_row=op["r2"], end_column=op["c2"])
        elif kind == "unmerge":
            ws.unmerge_cells(start_row=op["r1"], start_column=op["c1"],
                             end_row=op["r2"], end_column=op["c2"])
        else:
            log.warning(f"Unbekannte Operation übersprungen: {kind}")

    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()
